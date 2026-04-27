"""
Autonomous Volt - Pre-Autonomous Checklist + LUT (Web App)
===========================================================

Streamlit version of the desktop app. Runs in a browser, deployable to
Streamlit Community Cloud for a shareable URL.

Run locally:
    streamlit run app.py

Deploy:
    1) Push this folder to GitHub
    2) Go to https://share.streamlit.io -> New app -> point to app.py
    3) Share the URL with your team.

Logs are written to logs/pre_autonomous_checklist_log.xlsx in the working
directory. On Streamlit Cloud this resets on redeploy, so the app also
provides a "Download log file" button so runs can be saved locally.
"""

import os
from datetime import datetime, date
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from lut_module import (build_lut, predict_torque, run_controller_comparison,
                        analyze_real_data_files,
                        T_SPEC_CAP_LBFT, T_MAX_LBFT)


# ----------------------------
# Paths
# ----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_DIR  = os.path.join(BASE_DIR, "logs")
LOG_PATH = os.path.join(LOG_DIR, "pre_autonomous_checklist_log.xlsx")
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)


# ----------------------------
# Excel schema (same 3-tab architecture)
# ----------------------------
RUNS_HEADERS = ["RunID", "RunTimestamp", "Driver", "Date"]
CHECKLIST_PRE_HEADERS = [
    "RunID", "CalibrationDone",
    "TireFR_psi", "TireFL_psi", "TireRR_psi", "TireRL_psi",
    "Charge_pct", "InitialIncline", "Temperature_F", "Precipitation",
]
CHECKLIST_POST_HEADERS = [
    "RunID", "TargetSpeed_kph", "MaxSpeed_kph", "AnomalyObserved", "Notes", "RunOutcome",
    "DataQualityScore_0to5", "LUTCandidate", "DataFiles",
    "LUT_Baseline_lbft", "LUT_DeltaT_lbft", "LUT_TotalTorque_lbft",
    "LUT_OverSpecCap", "Outcome", "CompetitionReady",
]


# ----------------------------
# Helpers
# ----------------------------
def _sanitize_for_id(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return "unknown"
    return "".join(ch if (ch.isalnum() or ch in ("-", "_")) else "_" for ch in s)


def in_tire_range(x: float) -> bool:
    return 30.0 <= x <= 35.0


def ensure_log_workbook(path: str) -> None:
    def ensure_sheet(wb, title: str, headers: list) -> None:
        ws = wb[title] if title in wb.sheetnames else wb.create_sheet(title=title)
        for i, header in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=header)
            ws.column_dimensions[get_column_letter(i)].width = max(16, len(header) + 2)

    if os.path.exists(path):
        wb = load_workbook(path)
        ensure_sheet(wb, "Runs", RUNS_HEADERS)
        ensure_sheet(wb, "Checklist_Pre", CHECKLIST_PRE_HEADERS)
        ensure_sheet(wb, "Checklist_Post", CHECKLIST_POST_HEADERS)
        wb.save(path)
        return

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Runs"
    ws0.append(RUNS_HEADERS)
    for i, h in enumerate(RUNS_HEADERS, start=1):
        ws0.column_dimensions[get_column_letter(i)].width = max(16, len(h) + 2)

    ws_pre = wb.create_sheet("Checklist_Pre"); ws_pre.append(CHECKLIST_PRE_HEADERS)
    for i, h in enumerate(CHECKLIST_PRE_HEADERS, start=1):
        ws_pre.column_dimensions[get_column_letter(i)].width = max(16, len(h) + 2)

    ws_post = wb.create_sheet("Checklist_Post"); ws_post.append(CHECKLIST_POST_HEADERS)
    for i, h in enumerate(CHECKLIST_POST_HEADERS, start=1):
        ws_post.column_dimensions[get_column_letter(i)].width = max(16, len(h) + 2)

    wb.save(path)


def append_log_row_to_sheet(path: str, sheet_name: str, row: list) -> None:
    ensure_log_workbook(path)
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.insert_rows(2)
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=2, column=col_idx, value=value)
    wb.save(path)


def evaluate(inputs: dict) -> dict:
    calibration = inputs["calibration"]
    precip = inputs["precip"]
    fr, fl, rr, rl = (inputs["tire_fr"], inputs["tire_fl"],
                      inputs["tire_rr"], inputs["tire_rl"])
    charge = inputs["charge"]

    tires_ok = all(in_tire_range(v) for v in (fr, fl, rr, rl))
    autonomous_charge_ok = charge > 20

    if calibration == "Y" and tires_ok and autonomous_charge_ok:
        if charge > 50 and precip == "N":
            return {"outcome": "PASS", "detail": "COMPETITION READY",
                    "competition_ready": True, "light": "green"}
        return {"outcome": "PASS", "detail": "PASS — NOT COMPETITION READY",
                "competition_ready": False, "light": "yellow"}

    return {"outcome": "FAIL", "detail": "FAIL — NOT PRE-AUTONOMOUS READY",
            "competition_ready": False, "light": "red"}


# ============================================================
# Page setup & theme
# ============================================================
st.set_page_config(
    page_title="Autonomous Volt · Pre-Autonomous Checklist",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Custom CSS — automotive dashboard aesthetic to match desktop app
st.markdown("""
<style>
    /* Backgrounds */
    .stApp { background-color: #0f1418; }
    section[data-testid="stSidebar"] { background-color: #1a2028; }

    /* Text */
    .stApp, .stMarkdown, label, p, span, div { color: #e8ecef; }

    /* Headings in blue accent */
    h1, h2, h3 { color: #e8ecef !important; font-family: 'Segoe UI', sans-serif; }

    /* Input boxes */
    .stTextInput input, .stNumberInput input, .stTextArea textarea,
    .stSelectbox > div > div {
        background-color: #242c36 !important;
        color: #e8ecef !important;
        border: 1px solid #2d3640 !important;
        border-radius: 4px;
    }

    /* Radio buttons */
    .stRadio label { color: #e8ecef !important; }

    /* Primary buttons */
    .stButton > button {
        background-color: #4da3ff;
        color: white;
        font-weight: bold;
        border: none;
        padding: 10px 24px;
        border-radius: 4px;
        transition: background-color 0.15s;
    }
    .stButton > button:hover {
        background-color: #6bb4ff;
        color: white;
    }

    /* Download button */
    .stDownloadButton > button {
        background-color: #242c36;
        color: #e8ecef;
        border: 1px solid #2d3640;
    }

    /* Custom classes */
    .badge {
        display: inline-block;
        background: #2e7bc9;
        color: white;
        padding: 4px 12px;
        font-size: 10px;
        font-weight: bold;
        letter-spacing: 1px;
        border-radius: 2px;
    }
    .badge-muted {
        color: #6b7682;
        font-size: 10px;
        font-weight: bold;
        letter-spacing: 1px;
        margin-left: 8px;
    }
    .section-header {
        color: #4da3ff;
        font-size: 13px;
        font-weight: bold;
        letter-spacing: 1.5px;
        margin: 24px 0 8px 0;
        border-bottom: 1px solid #2d3640;
        padding-bottom: 6px;
    }
    .status-pill {
        display: inline-flex;
        align-items: center;
        background: #1a2028;
        padding: 8px 14px;
        border-radius: 4px;
        font-size: 10px;
        font-weight: bold;
        letter-spacing: 1px;
        color: #98a3ad;
    }
    .status-dot {
        display: inline-block;
        width: 10px;
        height: 10px;
        border-radius: 50%;
        margin-right: 8px;
    }
    .card {
        background: #1a2028;
        padding: 20px;
        border-radius: 6px;
        margin-bottom: 12px;
    }
    .card-alt {
        background: #242c36;
        padding: 20px;
        border-radius: 6px;
        margin-bottom: 12px;
    }
    .tire-label {
        color: #4da3ff;
        font-size: 10px;
        font-weight: bold;
        letter-spacing: 1px;
    }
    .result-status-green { color: #22c55e; font-size: 22px; font-weight: bold; }
    .result-status-yellow { color: #f59e0b; font-size: 22px; font-weight: bold; }
    .result-status-red { color: #ef4444; font-size: 22px; font-weight: bold; }
    .result-status-idle { color: #98a3ad; font-size: 18px; font-weight: bold; }
    .lamp {
        display: inline-block;
        width: 44px;
        height: 44px;
        border-radius: 50%;
        margin-right: 8px;
        background: #4a5662;
    }
    .lamp-red-on    { background: #ef4444; box-shadow: 0 0 20px #ef4444; }
    .lamp-yellow-on { background: #f59e0b; box-shadow: 0 0 20px #f59e0b; }
    .lamp-green-on  { background: #22c55e; box-shadow: 0 0 20px #22c55e; }

    .big-number {
        font-family: Consolas, 'Courier New', monospace;
        font-size: 28px;
        font-weight: bold;
    }
    .label-muted {
        color: #6b7682;
        font-size: 10px;
        font-weight: bold;
        letter-spacing: 1px;
    }
    .safety-banner-green {
        background: #14341f; color: #86efac;
        padding: 14px; text-align: center; font-weight: bold; border-radius: 4px;
    }
    .safety-banner-yellow {
        background: #3d2f1a; color: #fcd34d;
        padding: 14px; text-align: center; font-weight: bold; border-radius: 4px;
    }
    .safety-banner-red {
        background: #3d1f24; color: #fca5a5;
        padding: 14px; text-align: center; font-weight: bold; border-radius: 4px;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
# Cached LUT build (runs once, shared across user sessions)
# ============================================================
@st.cache_resource
def get_lut():
    return build_lut(data_folder=DATA_DIR, verbose=False)


@st.cache_resource
def get_real_data_analysis():
    return analyze_real_data_files(DATA_DIR)


lut = get_lut()
lut_source = getattr(lut, "data_source", "unknown")
lut_samples = getattr(lut, "n_samples", 0)


# ============================================================
# Header
# ============================================================
header_col1, header_col2 = st.columns([3, 1])

with header_col1:
    st.markdown(
        '<span class="badge">AUTONOMOUS VOLT</span>'
        '<span class="badge-muted">UW–MADISON CAPSTONE</span>',
        unsafe_allow_html=True
    )
    st.markdown("# Pre-Autonomous Checklist")
    st.markdown(
        '<p style="color:#98a3ad;">Verify vehicle state before every autonomous run  ·  '
        'results logged to Excel  ·  LUT predicts torque for run conditions</p>',
        unsafe_allow_html=True
    )

with header_col2:
    # Shorten source label ("CSV/XLSX from /long/path" -> "REAL CSV DATA")
    src_short = "REAL CSV DATA" if "CSV" in lut_source else "SYNTHETIC"
    st.markdown(
        f'<div style="text-align:right; padding-top:40px;">'
        f'<span class="status-pill">'
        f'<span class="status-dot" style="background:#22c55e;"></span>'
        f'LUT READY  ·  {src_short}  ·  {lut_samples:,} SAMPLES'
        f'</span></div>',
        unsafe_allow_html=True
    )

st.markdown("---")

# ============================================================
# LUT data diagnostics (collapsible)
# ============================================================
diagnostics = getattr(lut, "file_diagnostics", []) or []
if diagnostics:
    with st.expander(f"📊  LUT Training Data  ·  {len(diagnostics)} file(s)  ·  click to inspect"):
        st.markdown(
            '<p style="color:#98a3ad; font-size:12px;">'
            'These are the CSV/XLSX files in the <code>data/</code> folder that were '
            'loaded to train the LUT. Steering is converted from wheel angle to road-wheel '
            f'angle using a 15:1 ratio. Torque column is auto-converted N·m → lb-ft. '
            f'Rows outside the LUT envelope (speed ≤ 15 kph, steering ±30°) are dropped.'
            '</p>',
            unsafe_allow_html=True
        )

        rows = []
        for d in diagnostics:
            status_icon = {"ok": "✓", "read-error": "✗", "missing-cols": "⚠"}.get(
                d.get("status", "ok"), "?"
            )
            if d.get("status") == "ok":
                v_lo, v_hi = d.get("speed_range", (0, 0))
                s_lo, s_hi = d.get("steer_range", (0, 0))
                t_lo, t_hi = d.get("torque_range_lbft", (0, 0))
                rows.append({
                    "File": d["file"],
                    "Raw": d["raw_rows"],
                    "Valid": d["valid_rows"],
                    "Used": d.get("kept_rows", 0),
                    "Speed (kph)": f"{v_lo:.1f} – {v_hi:.1f}",
                    "Road steer (°)": f"{s_lo:+.1f} – {s_hi:+.1f}",
                    "Torque (lb-ft)": f"{t_lo:+.0f} – {t_hi:+.0f}",
                    "Notes": ", ".join(d.get("notes", [])),
                })
            else:
                rows.append({
                    "File": f"{status_icon} {d['file']}",
                    "Raw": d.get("raw_rows", 0),
                    "Valid": 0, "Used": 0,
                    "Speed (kph)": "—", "Road steer (°)": "—", "Torque (lb-ft)": "—",
                    "Notes": f"{d.get('status')}: {', '.join(d.get('notes', []))}",
                })

        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        total_raw = sum(d.get("raw_rows", 0) for d in diagnostics)
        total_used = sum(d.get("kept_rows", 0) for d in diagnostics)
        st.markdown(
            f'<p style="color:#6b7682; font-size:11px;">'
            f'Total: {total_raw:,} raw rows across {len(diagnostics)} file(s)  ·  '
            f'{total_used:,} rows used after envelope clipping  ·  '
            f'LUT has {lut_samples:,} training samples'
            f'</p>',
            unsafe_allow_html=True
        )

    st.markdown("")  # small spacer


# ============================================================
# Session state init
# ============================================================
if "last_result" not in st.session_state:
    st.session_state.last_result = None
if "awaiting_post_run" not in st.session_state:
    st.session_state.awaiting_post_run = False
if "pending_run_data" not in st.session_state:
    st.session_state.pending_run_data = None
if "lut_prediction" not in st.session_state:
    st.session_state.lut_prediction = None
if "form_nonce" not in st.session_state:
    st.session_state.form_nonce = 0
if "show_requery" not in st.session_state:
    st.session_state.show_requery = False
if "comparison_result" not in st.session_state:
    st.session_state.comparison_result = None


# ============================================================
# Top-level tabs
# ============================================================
tab_checklist, tab_compare, tab_realdata = st.tabs([
    "  PRE-AUTONOMOUS CHECKLIST  ",
    "  CONTROLLER COMPARISON  ",
    "  REAL DATA ANALYSIS  ",
])


with tab_checklist:


    # ============================================================
    # Main form (Pre-Run Checklist)
    # ============================================================
    with st.form(f"pre_checklist_{st.session_state.get('form_nonce', 0)}", clear_on_submit=False):
        st.markdown('<div class="section-header">SESSION</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            driver = st.text_input("Driver", placeholder="e.g. Patrick")
        with c2:
            run_date = st.text_input("Date (YYYY-MM-DD)", value=date.today().isoformat())

        st.markdown('<div class="section-header">VEHICLE STATE</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            calib = st.radio("Calibration complete", ["Yes", "No"], index=1, horizontal=True)
        with c2:
            charge = st.number_input("Charge (%)", min_value=0, max_value=100, value=0, step=1,
                                     help="> 20% PASS  ·  > 50% COMP READY")

        st.markdown('<div class="section-header">ENVIRONMENT</div>', unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            precip = st.radio("Precipitation", ["Yes", "No"], index=1, horizontal=True)
        with c2:
            incline = st.number_input("Initial incline (%)", min_value=-20.0, max_value=20.0,
                                      value=0.0, step=0.5)
        with c3:
            temperature = st.number_input("Temperature (°F)", min_value=-20, max_value=120,
                                          value=70, step=1)

        st.markdown(
            '<div class="section-header">TIRE PRESSURE · 30–35 psi</div>',
            unsafe_allow_html=True
        )
        t1, t2, t3, t4 = st.columns(4)
        with t1:
            st.markdown('<div class="tire-label">FR</div>', unsafe_allow_html=True)
            tire_fr = st.number_input("FR (psi)", min_value=0.0, max_value=60.0,
                                       value=32.0, step=0.5, label_visibility="collapsed")
        with t2:
            st.markdown('<div class="tire-label">FL</div>', unsafe_allow_html=True)
            tire_fl = st.number_input("FL (psi)", min_value=0.0, max_value=60.0,
                                       value=32.0, step=0.5, label_visibility="collapsed")
        with t3:
            st.markdown('<div class="tire-label">RR</div>', unsafe_allow_html=True)
            tire_rr = st.number_input("RR (psi)", min_value=0.0, max_value=60.0,
                                       value=32.0, step=0.5, label_visibility="collapsed")
        with t4:
            st.markdown('<div class="tire-label">RL</div>', unsafe_allow_html=True)
            tire_rl = st.number_input("RL (psi)", min_value=0.0, max_value=60.0,
                                       value=32.0, step=0.5, label_visibility="collapsed")

        submitted = st.form_submit_button("▶  RUN CHECKLIST", use_container_width=False)


    # ============================================================
    # Evaluate the checklist when submitted
    # ============================================================
    if submitted:
        if not driver.strip():
            st.error("Please enter a Driver name.")
        elif not run_date.strip():
            st.error("Please enter a Date.")
        else:
            inputs = {
                "driver": driver.strip(), "date": run_date.strip(),
                "calibration": "Y" if calib == "Yes" else "N",
                "tire_fr": tire_fr, "tire_fl": tire_fl,
                "tire_rr": tire_rr, "tire_rl": tire_rl,
                "charge": int(charge),
                "incline": str(incline),
                "weather": str(temperature),
                "precip": "Y" if precip == "Yes" else "N",
            }
            result = evaluate(inputs)
            st.session_state.last_result = result
            st.session_state.pending_run_data = inputs

            if result["competition_ready"]:
                st.session_state.awaiting_post_run = True
                st.session_state.lut_prediction = None
            else:
                # Log FAIL / not-comp-ready immediately
                now = datetime.now()
                timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
                run_id = f"{now.strftime('%Y%m%d_%H%M%S')}_{_sanitize_for_id(driver)}"

                try:
                    append_log_row_to_sheet(LOG_PATH, "Runs",
                                            [run_id, timestamp, inputs["driver"], inputs["date"]])
                    append_log_row_to_sheet(LOG_PATH, "Checklist_Pre", [
                        run_id, inputs["calibration"],
                        inputs["tire_fr"], inputs["tire_fl"],
                        inputs["tire_rr"], inputs["tire_rl"],
                        inputs["charge"], inputs["incline"],
                        inputs["weather"], inputs["precip"],
                    ])
                    append_log_row_to_sheet(LOG_PATH, "Checklist_Post", [
                        run_id, "", "", "", "", "", "", "", "",
                        "", "", "", "",
                        result["outcome"], "N",
                    ])
                    st.session_state.logged_run_id = run_id
                except Exception as e:
                    st.warning(f"Run evaluated, but log write failed: {e}")
                st.session_state.awaiting_post_run = False


    # ============================================================
    # Result display (always shown after a run)
    # ============================================================
    if st.session_state.last_result is not None:
        result = st.session_state.last_result

        st.markdown('<div class="section-header">RESULT</div>', unsafe_allow_html=True)

        with st.container():
            st.markdown('<div class="card-alt">', unsafe_allow_html=True)
            rc1, rc2 = st.columns([1, 3])
            with rc1:
                # Traffic light
                light = result["light"]
                red    = "lamp-red-on"    if light == "red"    else ""
                yellow = "lamp-yellow-on" if light == "yellow" else ""
                green  = "lamp-green-on"  if light == "green"  else ""
                st.markdown(
                    f'<div style="padding-top:10px;">'
                    f'<span class="lamp {red}"></span>'
                    f'<span class="lamp {yellow}"></span>'
                    f'<span class="lamp {green}"></span>'
                    f'</div>',
                    unsafe_allow_html=True
                )

            with rc2:
                st.markdown(
                    '<div class="label-muted">STATUS</div>',
                    unsafe_allow_html=True
                )
                status_class = {"green": "result-status-green",
                                "yellow": "result-status-yellow",
                                "red": "result-status-red"}[light]
                st.markdown(
                    f'<div class="{status_class}">{result["detail"]}</div>',
                    unsafe_allow_html=True
                )

                inputs = st.session_state.pending_run_data
                tires_ok = all(in_tire_range(v) for v in
                               (inputs["tire_fr"], inputs["tire_fl"],
                                inputs["tire_rr"], inputs["tire_rl"]))
                detail = (
                    f'Tires {inputs["tire_fr"]:.1f} / {inputs["tire_fl"]:.1f} / '
                    f'{inputs["tire_rr"]:.1f} / {inputs["tire_rl"]:.1f} psi — '
                    f'{"OK" if tires_ok else "OUT OF RANGE"}   ·   '
                    f'Calibration: {inputs["calibration"]}   ·   '
                    f'Charge: {inputs["charge"]}%   ·   '
                    f'Precip: {inputs["precip"]}   ·   '
                    f'Temp: {inputs["weather"]}°F   ·   '
                    f'Incline: {inputs["incline"]}%'
                )
                st.markdown(
                    f'<div style="color:#6b7682; font-size:11px; margin-top:6px;">{detail}</div>',
                    unsafe_allow_html=True
                )
            st.markdown('</div>', unsafe_allow_html=True)


    # ============================================================
    # Post-Run Checklist (shown only if competition_ready)
    # ============================================================
    if st.session_state.awaiting_post_run:
        st.markdown("---")
        st.markdown('<span class="badge">POST-RUN CHECKLIST</span>', unsafe_allow_html=True)
        st.markdown("## Capture results from the autonomous run")

        with st.form("post_run", clear_on_submit=False):
            c1, c2 = st.columns(2)
            with c1:
                target_speed = st.number_input(
                    "Target speed (for LUT) [kph]",
                    min_value=0.0, max_value=50.0, value=8.0, step=0.5,
                    help="Setpoint the controller was holding · feeds the LUT query"
                )
            with c2:
                max_speed = st.number_input(
                    "Max speed reached [kph]",
                    min_value=0.0, max_value=50.0, value=10.0, step=0.5,
                    help="Peak speed observed during run · logged only"
                )

            c1, c2 = st.columns(2)
            with c1:
                anomaly = st.radio("Anomaly observed", ["Yes", "No"], index=1, horizontal=True)
            with c2:
                run_outcome = st.selectbox("Run outcome",
                                            ["Completed", "Aborted", "SafetyStop"], index=0)

            notes = st.text_input("Notes", placeholder="Optional driver notes")

            c1, c2, c3 = st.columns(3)
            with c1:
                data_quality = st.number_input("Data quality (0–5)", min_value=0, max_value=5,
                                                value=3, step=1)
            with c2:
                lut_candidate = st.radio("LUT candidate", ["Yes", "No"], index=0, horizontal=True)
            with c3:
                data_files = st.text_input("Data file(s)", placeholder="e.g. run_042.csv")

            with st.expander("📋  LUT Data Quality Reference"):
                st.markdown("""
                | Parameter | Limit | Priority |
                |---|---|---|
                | Vehicle Speed | < 15 kph | **MUST** |
                | Grade Envelope | ≤ 10 % | Nice |
                | Deceleration Limit | ≤ 3 m/s² | **MUST** |
                | Steering Rate Limit | ≤ 15 deg/s | **MUST** |
                | Maximum Torque Command | ≤ 185 lbf-ft | **MUST** |
                """)

            save_post = st.form_submit_button("SAVE & PREDICT TORQUE", use_container_width=False)

        if save_post:
            inputs = st.session_state.pending_run_data

            # Query LUT at target speed (the feedforward design point)
            try:
                grade = float(inputs["incline"]) if inputs["incline"] else 0.0
            except ValueError:
                grade = 0.0

            pred = predict_torque(lut, target_speed, grade, 0.0)

            # Log to Excel
            now = datetime.now()
            timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
            run_id = f"{now.strftime('%Y%m%d_%H%M%S')}_{_sanitize_for_id(inputs['driver'])}"

            try:
                append_log_row_to_sheet(LOG_PATH, "Runs",
                                        [run_id, timestamp, inputs["driver"], inputs["date"]])
                append_log_row_to_sheet(LOG_PATH, "Checklist_Pre", [
                    run_id, inputs["calibration"],
                    inputs["tire_fr"], inputs["tire_fl"],
                    inputs["tire_rr"], inputs["tire_rl"],
                    inputs["charge"], inputs["incline"],
                    inputs["weather"], inputs["precip"],
                ])
                append_log_row_to_sheet(LOG_PATH, "Checklist_Post", [
                    run_id, target_speed, max_speed,
                    "Y" if anomaly == "Yes" else "N",
                    notes, run_outcome, data_quality,
                    "Y" if lut_candidate == "Yes" else "N",
                    data_files,
                    round(pred["baseline_lbft"], 2),
                    round(pred["delta_lbft"], 2),
                    round(pred["total_lbft"], 2),
                    "Y" if pred["over_spec_cap"] else "N",
                    st.session_state.last_result["outcome"], "Y",
                ])
                st.session_state.logged_run_id = run_id
            except Exception as e:
                st.warning(f"Log write failed: {e}")

            st.session_state.lut_prediction = {
                "target_speed": target_speed,
                "grade": grade,
                "steer": 0.0,
                "pred": pred,
            }
            st.session_state.awaiting_post_run = False
            st.rerun()


    # ============================================================
    # LUT Prediction display
    # ============================================================
    if st.session_state.lut_prediction is not None:
        st.markdown("---")
        p = st.session_state.lut_prediction
        pred = p["pred"]

        st.markdown('<span class="badge">LOOKUP TABLE PREDICTION</span>', unsafe_allow_html=True)
        st.markdown("## Torque forecast for run conditions")

        # Inputs card
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="label-muted">QUERY INPUTS</div>', unsafe_allow_html=True)
        i1, i2, i3 = st.columns(3)
        with i1:
            st.markdown(f'**Target speed (setpoint)**<br>{p["target_speed"]:.2f} kph',
                        unsafe_allow_html=True)
        with i2:
            st.markdown(f'**Grade (initial incline)**<br>{p["grade"]:.2f} %',
                        unsafe_allow_html=True)
        with i3:
            st.markdown(f'**Steering angle (assumed)**<br>{p["steer"]:.2f} deg',
                        unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        # Predicted torque - three big numbers
        total = pred["total_lbft"]
        if pred["over_spec_cap"]:
            total_color = "#ef4444"
        elif abs(total) > 0.8 * T_SPEC_CAP_LBFT:
            total_color = "#f59e0b"
        else:
            total_color = "#22c55e"

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="label-muted">PREDICTED TORQUE</div>', unsafe_allow_html=True)
        b1, b2, b3 = st.columns(3)
        with b1:
            st.markdown(
                f'<div class="label-muted">BASELINE</div>'
                f'<div class="big-number" style="color:#e8ecef;">{pred["baseline_lbft"]:+.2f}</div>'
                f'<div style="color:#6b7682; font-size:11px;">lb-ft</div>',
                unsafe_allow_html=True
            )
        with b2:
            st.markdown(
                f'<div class="label-muted">Δ FROM LUT</div>'
                f'<div class="big-number" style="color:#4da3ff;">{pred["delta_lbft"]:+.2f}</div>'
                f'<div style="color:#6b7682; font-size:11px;">lb-ft</div>',
                unsafe_allow_html=True
            )
        with b3:
            st.markdown(
                f'<div class="label-muted">TOTAL</div>'
                f'<div class="big-number" style="color:{total_color};">{total:+.2f}</div>'
                f'<div style="color:#6b7682; font-size:11px;">lb-ft</div>',
                unsafe_allow_html=True
            )
        st.markdown('</div>', unsafe_allow_html=True)

        # Safety banner
        if pred["over_spec_cap"]:
            st.markdown(
                f'<div class="safety-banner-red">'
                f'⚠  EXCEEDS SPEC CAP  ·  ±{T_SPEC_CAP_LBFT:.0f} lb-ft  ·  REVIEW CALIBRATION'
                f'</div>',
                unsafe_allow_html=True
            )
        elif abs(total) > 0.8 * T_SPEC_CAP_LBFT:
            st.markdown(
                '<div class="safety-banner-yellow">'
                '⚠  WITHIN 80% OF SPEC CAP  ·  PROCEED WITH CAUTION'
                '</div>',
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                f'<div class="safety-banner-green">'
                f'✓  WITHIN SPEC ENVELOPE  ·  ±{T_SPEC_CAP_LBFT:.0f} lb-ft'
                f'</div>',
                unsafe_allow_html=True
            )

        st.markdown(
            f'<div style="color:#6b7682; font-size:11px; margin-top:16px;">'
            f'LUT source: {lut_source}  ·  {lut_samples:,} samples  ·  '
            f'motor cap ±{T_MAX_LBFT:.0f} lb-ft'
            f'</div>',
            unsafe_allow_html=True
        )

        if hasattr(st.session_state, "logged_run_id"):
            st.success(f"✓  Logged to Excel. RunID: `{st.session_state.logged_run_id}`")

        # ---------- Next-action buttons ----------
        st.markdown("<br>", unsafe_allow_html=True)
        nb1, nb2, nb3 = st.columns([1, 1, 2])
        with nb1:
            if st.button("▶  NEW CHECKLIST RUN", use_container_width=True, type="primary"):
                # Full reset - clear all session state except the cached LUT
                for key in ["last_result", "awaiting_post_run", "pending_run_data",
                            "lut_prediction", "logged_run_id", "show_requery"]:
                    if key in st.session_state:
                        del st.session_state[key]
                # Bump the form key so inputs reset to defaults
                st.session_state.form_nonce = st.session_state.get("form_nonce", 0) + 1
                st.rerun()
        with nb2:
            if st.button("🔁  PREDICT ANOTHER", use_container_width=True):
                # Keep the checklist inputs, just reopen the torque query
                st.session_state.show_requery = True
                st.rerun()


    # ============================================================
    # Re-query panel - lets the user predict torque at new conditions
    # without re-filling the full checklist
    # ============================================================
    if st.session_state.get("show_requery"):
        st.markdown("---")
        st.markdown('<span class="badge">NEW TORQUE QUERY</span>', unsafe_allow_html=True)
        st.markdown("## Predict torque at different conditions")
        st.markdown(
            '<p style="color:#98a3ad;">Adjust the operating point and re-query the LUT. '
            'This does not change the checklist or add a new log entry.</p>',
            unsafe_allow_html=True
        )

        with st.form("requery_form", clear_on_submit=False):
            r1, r2, r3 = st.columns(3)
            with r1:
                rq_speed = st.number_input(
                    "Target speed [kph]",
                    min_value=0.0, max_value=50.0, value=8.0, step=0.5,
                )
            with r2:
                rq_grade = st.number_input(
                    "Grade [%]",
                    min_value=-20.0, max_value=20.0, value=0.0, step=0.5,
                )
            with r3:
                rq_steer = st.number_input(
                    "Road-wheel steering angle [deg]",
                    min_value=-31.0, max_value=31.0, value=0.0, step=1.0,
                    help="Road-wheel angle (not steering wheel). "
                         "±30° ≈ full lock on the Volt at 15:1 ratio."
                )

            rq_col1, rq_col2 = st.columns([1, 1])
            with rq_col1:
                rq_submit = st.form_submit_button("QUERY LUT", use_container_width=True,
                                                   type="primary")
            with rq_col2:
                rq_cancel = st.form_submit_button("CANCEL", use_container_width=True)

        if rq_submit:
            pred = predict_torque(lut, rq_speed, rq_grade, rq_steer)
            st.session_state.lut_prediction = {
                "target_speed": rq_speed,
                "grade": rq_grade,
                "steer": rq_steer,
                "pred": pred,
            }
            # Clear the "logged_run_id" message since this isn't a new logged run
            if "logged_run_id" in st.session_state:
                del st.session_state["logged_run_id"]
            st.session_state.show_requery = False
            st.rerun()

        if rq_cancel:
            st.session_state.show_requery = False
            st.rerun()


    # ============================================================
    # Log download section (always available at bottom)
    # ============================================================
    st.markdown("---")
    dl_col1, dl_col2 = st.columns([3, 1])
    with dl_col1:
        st.markdown("**Excel Log**  ·  all runs with RunID linking across 3 tabs")
        if os.path.exists(LOG_PATH):
            try:
                # Show quick stats
                wb = load_workbook(LOG_PATH, read_only=True)
                n_runs = wb["Runs"].max_row - 1
                st.markdown(
                    f'<span style="color:#6b7682;">{n_runs} run(s) logged  ·  '
                    f'3 tabs: Runs, Checklist_Pre, Checklist_Post</span>',
                    unsafe_allow_html=True
                )
            except Exception:
                pass

    with dl_col2:
        if os.path.exists(LOG_PATH):
            with open(LOG_PATH, "rb") as f:
                st.download_button(
                    label="📥  DOWNLOAD LOG",
                    data=f.read(),
                    file_name="pre_autonomous_checklist_log.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
        else:
            st.button("No log yet", disabled=True, use_container_width=True)


with tab_compare:
    st.markdown('<span class="badge">CONTROLLER COMPARISON</span>',
                unsafe_allow_html=True)
    st.markdown("# Baseline PID vs. LUT-based Controller")
    st.markdown(
        '<p style="color:#98a3ad;">'
        '30-second closed-loop simulation comparing the existing PID controller '
        '(no lookup table) against the LUT-based feedforward + PI controller. '
        'Both run through the same drive cycle: speed setpoint sweeping 6–10 kph, '
        'grade ranging ±6%, steering up to ±20°.'
        '</p>',
        unsafe_allow_html=True
    )

    # Data source selector
    src_col1, src_col2 = st.columns([2, 1])
    with src_col1:
        sim_source = st.radio(
            "LUT data source for simulation",
            options=["Synthetic (matches simulation physics)",
                     "Real CSV data (your team’s tests)"],
            index=0,
            horizontal=True,
            help="The simulation uses an idealized bicycle model. Synthetic training "
                 "data was generated from that same model, so the LUT’s corrections "
                 "match the simulation exactly. Real CSV data captures actual vehicle "
                 "behavior including effects the simulation does not model."
        )
    with src_col2:
        st.markdown("<br>", unsafe_allow_html=True)
        run_sim = st.button("▶  RUN SIMULATION", use_container_width=True,
                            type="primary")

    if run_sim:
        with st.spinner("Running 30-second drive cycle on both controllers…"):
            force_syn = sim_source.startswith("Synthetic")
            sim_lut = build_lut(data_folder=DATA_DIR, verbose=False,
                                force_synthetic=force_syn)
            comp = run_controller_comparison(sim_lut, duration_s=30.0)
            st.session_state.comparison_result = {
                "comp": comp,
                "source_label": sim_source,
                "n_samples": getattr(sim_lut, "n_samples", 0),
            }

    if st.session_state.comparison_result is not None:
        comp = st.session_state.comparison_result["comp"]
        src_label = st.session_state.comparison_result["source_label"]
        n_samples = st.session_state.comparison_result["n_samples"]

        bm = comp["baseline"]["metrics"]
        am = comp["adaptive"]["metrics"]
        imp = comp["improvement"]

        st.markdown(
            f'<p style="color:#6b7682; font-size:11px;">'
            f'LUT source: {src_label}  ·  {n_samples:,} samples  ·  '
            f'30 s cycle  ·  600 timesteps'
            f'</p>',
            unsafe_allow_html=True
        )

        # ----- Improvement summary cards (top) -----
        st.markdown('<div class="section-header">IMPROVEMENT SUMMARY</div>',
                    unsafe_allow_html=True)

        i1, i2, i3, i4 = st.columns(4)
        def imp_card(col, label, value, suffix, good_if_positive=True):
            color = "#22c55e" if (value > 0) == good_if_positive else "#ef4444"
            with col:
                st.markdown(
                    f'<div class="card">'
                    f'<div class="label-muted">{label}</div>'
                    f'<div class="big-number" style="color:{color};">{value:+.1f}</div>'
                    f'<div style="color:#6b7682; font-size:11px;">{suffix}</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )

        imp_card(i1, "IN-BAND TIME", imp["inband_delta_pct"], "percentage points")
        imp_card(i2, "RMSE SPEED", imp["rmse_v_change_pct"], "% reduction")
        imp_card(i3, "RMS JERK", imp["rms_jerk_change_pct"], "% reduction")
        imp_card(i4, "TORQUE SLEW", imp["torque_slew_change_pct"], "% reduction")

        # ----- Side-by-side metrics tables -----
        st.markdown('<div class="section-header">PER-CONTROLLER METRICS</div>',
                    unsafe_allow_html=True)

        tcol1, tcol2 = st.columns(2)
        rows_template = [
            ("Speed within ±0.5 kph",   "inband_pct",                "%",    1),
            ("RMSE speed error",              "rmse_v_kph",                "kph",  3),
            ("Max speed error",               "max_err_kph",               "kph",  3),
            ("RMS acceleration",              "rms_accel_mps2",            "m/s²",  4),
            ("RMS jerk",                      "rms_jerk_mps3",             "m/s³",  3),
            ("95th percentile |jerk|",        "p95_jerk_mps3",             "m/s³",  3),
            ("RMS torque slew rate",          "rms_torque_slew_lbft_per_s", "lb-ft/s", 1),
        ]

        def metric_table(col, title, m, color):
            with col:
                st.markdown(
                    f'<div class="card">'
                    f'<div class="label-muted" style="color:{color};">{title}</div>',
                    unsafe_allow_html=True
                )
                rows_html = ['<table style="width:100%; margin-top:10px; '
                             'color:#e8ecef; font-size:13px;">']
                for label, key, unit, dec in rows_template:
                    val = m[key]
                    rows_html.append(
                        f'<tr>'
                        f'<td style="padding:4px 0; color:#98a3ad;">{label}</td>'
                        f'<td style="padding:4px 0; text-align:right; '
                        f'font-family:Consolas,monospace;">{val:.{dec}f} {unit}</td>'
                        f'</tr>'
                    )
                rows_html.append("</table></div>")
                st.markdown("\n".join(rows_html), unsafe_allow_html=True)

        metric_table(tcol1, "BASELINE (PID only)",      bm, "#ef4444")
        metric_table(tcol2, "ADAPTIVE (PID + LUT)",     am, "#22c55e")

        # ----- Time-series plots -----
        st.markdown('<div class="section-header">TIME-SERIES TRACES</div>',
                    unsafe_allow_html=True)

        base_h = comp["baseline"]["history"]
        adap_h = comp["adaptive"]["history"]
        cols = comp["history_columns"]

        # Build comparison dataframes for st.line_chart
        df_speed = pd.DataFrame({
            "Target":    base_h[:, 2],
            "Baseline":  base_h[:, 1],
            "Adaptive":  adap_h[:, 1],
        }, index=base_h[:, 0])
        df_speed.index.name = "Time (s)"

        df_torque = pd.DataFrame({
            "Baseline":  base_h[:, 5],
            "Adaptive":  adap_h[:, 5],
        }, index=base_h[:, 0])
        df_torque.index.name = "Time (s)"

        df_jerk = pd.DataFrame({
            "Baseline":  base_h[:, 7],
            "Adaptive":  adap_h[:, 7],
        }, index=base_h[:, 0])
        df_jerk.index.name = "Time (s)"

        st.markdown("**Vehicle speed [kph]** — actual vs. target. Closer to target is better.")
        st.line_chart(df_speed, height=280, use_container_width=True)

        st.markdown("**Torque command [lb-ft]** — smoother is better (less mechanical wear, less jerky motion).")
        st.line_chart(df_torque, height=240, use_container_width=True)

        st.markdown("**Jerk [m/s³]** — lower magnitude means smoother ride.")
        st.line_chart(df_jerk, height=200, use_container_width=True)

    else:
        st.info("Click **RUN SIMULATION** above to compare the two controllers.")


with tab_realdata:
    st.markdown('<span class="badge">REAL DATA ANALYSIS</span>', unsafe_allow_html=True)
    st.markdown("# Jerk and Torque Slew from Recorded Tests")
    st.markdown(
        '<p style="color:#98a3ad;">'
        'Computed directly from your team\u2019s test recordings. Speed is differentiated '
        'twice (with Savitzky-Golay smoothing in between) to estimate acceleration and '
        'jerk; torque is differentiated once for slew rate. Stats use the middle 95% of '
        'each recording to ignore startup transients.'
        '</p>',
        unsafe_allow_html=True
    )

    rda = get_real_data_analysis()
    if not rda.get("files"):
        st.warning("No CSV files found in the data folder.")
    else:
        agg = rda["aggregate"]
        files = rda["files"]

        # Aggregate cards
        st.markdown('<div class="section-header">AGGREGATE \u00B7 ALL FILES</div>',
                    unsafe_allow_html=True)
        a1, a2, a3, a4 = st.columns(4)
        def agg_card(col, label, value, suffix):
            with col:
                st.markdown(
                    f'<div class="card">'
                    f'<div class="label-muted">{label}</div>'
                    f'<div class="big-number" style="color:#e8ecef;">{value}</div>'
                    f'<div style="color:#6b7682; font-size:11px;">{suffix}</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )
        agg_card(a1, "FILES ANALYZED", f'{agg["n_files"]}', "CSV/XLSX")
        agg_card(a2, "TOTAL SAMPLES", f'{agg["total_samples"]:,}', "raw rows")
        agg_card(a3, "MEDIAN RMS JERK", f'{agg["median_rms_jerk"]:.2f}', "m/s\u00B3")
        agg_card(a4, "MEDIAN TORQUE SLEW", f'{agg["median_rms_slew"]:.0f}', "lb-ft/s")

        # Per-file table
        st.markdown('<div class="section-header">PER-FILE STATS</div>',
                    unsafe_allow_html=True)
        df_table = pd.DataFrame([{
            "File":             f["name"],
            "Samples":          f["n_samples"],
            "Duration (s)":     round(f["duration_s"], 1),
            "Mean speed (kph)": round(f["mean_speed_kph"], 2),
            "RMS jerk (m/s\u00B3)":   round(f["rms_jerk_mps3"], 3),
            "P95 jerk (m/s\u00B3)":   round(f["p95_jerk_mps3"], 3),
            "RMS slew (lb-ft/s)": round(f["rms_torque_slew_lbft_per_s"], 1),
            "Max torque (lb-ft)": round(f["max_torque_lbft"], 0),
        } for f in files])
        st.dataframe(df_table, use_container_width=True, hide_index=True, height=420)

        # Per-file inspector
        st.markdown('<div class="section-header">FILE INSPECTOR</div>',
                    unsafe_allow_html=True)
        st.markdown(
            '<p style="color:#98a3ad; font-size:12px;">'
            'Pick a file to see its time-series traces.'
            '</p>',
            unsafe_allow_html=True
        )
        names = [f["name"] for f in files]
        # Default to a smooth file (5 mph constant speed) if available
        default_idx = next((i for i, n in enumerate(names)
                            if "5mph_straightline" in n or "5mph straightline" in n), 0)
        chosen = st.selectbox("File", names, index=default_idx, label_visibility="collapsed")

        sel = next(f for f in files if f["name"] == chosen)
        tr = sel["_traces"]

        # Summary line for the chosen file
        st.markdown(
            f'<p style="color:#6b7682; font-size:12px;">'
            f'<b>{sel["name"]}</b>  \u00B7  '
            f'{sel["n_samples"]:,} samples  \u00B7  '
            f'{sel["duration_s"]:.1f} s  \u00B7  '
            f'mean speed {sel["mean_speed_kph"]:.2f} kph  \u00B7  '
            f'RMS jerk {sel["rms_jerk_mps3"]:.3f} m/s\u00B3  \u00B7  '
            f'RMS slew {sel["rms_torque_slew_lbft_per_s"]:.1f} lb-ft/s'
            f'</p>',
            unsafe_allow_html=True
        )

        # Traces
        df_v = pd.DataFrame({"Speed (kph)": tr["v_kph"]}, index=tr["t"])
        df_v.index.name = "Time (s)"
        df_T = pd.DataFrame({"Torque (lb-ft)": tr["T_lbft"]}, index=tr["t"])
        df_T.index.name = "Time (s)"
        df_j = pd.DataFrame({"Jerk (m/s\u00B3)": tr["jerk"]}, index=tr["t"])
        df_j.index.name = "Time (s)"

        st.markdown("**Vehicle speed**")
        st.line_chart(df_v, height=220, use_container_width=True)
        st.markdown("**Axle torque**")
        st.line_chart(df_T, height=220, use_container_width=True)
        st.markdown("**Computed jerk** (filtered)")
        st.line_chart(df_j, height=220, use_container_width=True)

        with st.expander("\u2139\ufe0f  About these numbers"):
            st.markdown(
                """
                **Method**

                - Speed is filtered with a Savitzky-Golay filter (window 21, order 3),
                  differentiated to acceleration, filtered again, and differentiated
                  to jerk. Two derivatives amplify noise, so smoothing is necessary.
                - Torque is filtered the same way, then differentiated once for slew.
                - Stats are computed on the middle 95% of each recording to remove
                  startup and shutdown transients.

                **What these numbers mean**

                - **RMS jerk** captures average smoothness across the full recording.
                  Lower is smoother.
                - **P95 jerk** (95th percentile) captures the worst typical jerk
                  spikes during the run, while ignoring extreme outliers.
                - **RMS torque slew** measures how aggressively the throttle was
                  being applied or released.

                **Limitations**

                - These are derived from speed, not from an inertial accelerometer.
                  Numbers are order-of-magnitude estimates, not lab-grade
                  measurements.
                - The Bolt's IMU was not captured in these CSVs. Future testing
                  should record IMU data alongside torque and steering for
                  ground-truth jerk.
                - The 12 mph file mostly operates above the 15 kph LUT envelope, so
                  its high jerk values reflect operation outside the design
                  specification rather than poor LUT performance.
                """
            )
